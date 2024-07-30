# Importei as bibliotecas
import requests
import openpyxl
import json

# # Informei para o usuario como deve colocado o CNPJ
# print("O CNPJ deve conter 14 números, pode ter caracteres especiais:")

# # Criei um looping
# while True:
#     # Pedi para o Usuario digitar o CNPJ ou sair da aplicação
#     cn = input("Digite o CNPJ: ")
#     # Aqui e caso ele aperte sair
#     if cn.lower() == "sair":
#         break
#     # Estou fazendo a formatação do CNPJ para atender o jeito que a API pede
#     cn = cn.replace("-", "").replace(".", "").replace("/", "")

#     # Defeni o limite de caracteres do cnpj falando oq ele faz se for 14 numeros ou mias ou menos
#     if len(cn) == 14:
#         # Ele acessa a API com o token
#         #link = f'https://api.datastone.com.br/v1/companies/?cnpj={cn}&fields=all'
#         #headers = {"Authorization": "Token 2a466789-1dc2-491f-9098-cbc99dcbed38"}

#         # Pega o json fornecido da API
#         data = requests.get(link, headers=headers)

with open("empresascnpjPALIN&MARTINS.json") as arq:
    data = json.load(arq)
        
        # Defeni o que ele faz se der aceito a API
    if data.status_code == 200:
            # Defeni daonde ele vai importar as informações
            company_data = data.json()[0]
            
            # Aqui fiz ele pegar as informações que irei utilizar            
            razao_social = company_data.get('company_name')
            endereco = company_data.get('addresses')[0]
            cidade = company_data.get('city')
            estado = company_data.get('district')
            cnae = company_data.get('cnae_description')
            data_abertura = company_data.get('creation_date')
            emails = [email['email'] for email in company_data.get('emails')]
            filial_or_matriz = company_data.get('headquarter_type')
            telefones_fixo = [f"{tel['ddd']} {tel['number']}" for tel in company_data.get('land_lines')]
            celulares = [f"{cel['ddd']} {cel['number']}" for cel in company_data.get('mobile_phones')]
            socios = [socio['name'] for socio in company_data.get('related_persons')]
            
            # Criar ou acessar a planilha, depende.
            try:
                wb = openpyxl.load_workbook('AiimsColetados2.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                
            # Aqui ele mantem a planilha ativa para não criar varias
            sheet = wb.active
            linha_planilha = sheet.max_row + 1
            
            # Poem as informações na planilha
            sheet.cell(row=linha_planilha, column=4).value = razao_social
            sheet.cell(row=linha_planilha, column=5).value = cn
            sheet.cell(row=linha_planilha, column=6).value = cidade
            sheet.cell(row=linha_planilha, column=7).value = estado
            sheet.cell(row=linha_planilha, column=8).value = celulares
            sheet.cell(row=linha_planilha, column=9).value = emails
            sheet.cell(row=linha_planilha, column=11).value = cnae
            sheet.cell(row=linha_planilha, column=12).value = filial_or_matriz
            sheet.cell(row=linha_planilha, column=13).value = ', '.join(socios)
            
            
            # Salva a planilha
            wb.save('AiimsColetados2.xlsx')
             # Informa que salvou as informações da empresa
            print("Informações salvas com sucesso!")
            
            # Aqui é para falar se deu erro
    else:
            print("Erro ao acessar o CNPJ:", data.status_code)
else:
print("CNPJ Inválido")

    # Aqui para informar o fim
    print("Processo finalizado")