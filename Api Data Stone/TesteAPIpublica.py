import tkinter as tk
from tkinter import messagebox
import requests
import openpyxl

def consultar_cnpj():
    cn = entry_cnpj.get().replace("-", "").replace(".", "").replace("/", "")
    

    if len(cn) == 14:
        link = f'https://publica.cnpj.ws/cnpj/{cn}'
        requisicao = requests.get(link)
        
        if requisicao.status_code == 200:
            dic_requisicao = requisicao.json()
            razao_social = dic_requisicao['razao_social']
            cidade = dic_requisicao['estabelecimento']['cidade']['nome']
            estado = dic_requisicao['estabelecimento']['estado']['nome']
            ddd1 = dic_requisicao['estabelecimento']['ddd1']
            ddd2 = dic_requisicao['estabelecimento']['ddd2']
            telefone1 = dic_requisicao['estabelecimento']['telefone1']
            telefone2 = dic_requisicao['estabelecimento']['telefone2']
            #capital_social = dic_requisicao['capital_social']
            email = dic_requisicao['estabelecimento']['email']
            #socios = [socio['nome'] for socio in dic_requisicao['socios']]
            #socios_str = ', '.join(socios)
            
            try:
                wb = openpyxl.load_workbook('AiimsColetados2.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                
            sheet = wb.active
            linha_planilha = sheet.max_row + 1
            
            sheet.cell(row=linha_planilha, column=1).value = remove_caracter(razao_social)
            sheet.cell(row=linha_planilha, column=2).value = cnpj(cn)
            sheet.cell(row=linha_planilha, column=3).value = cidade
            sheet.cell(row=linha_planilha, column=4).value = estado
            sheet.cell(row=linha_planilha, column=5).value = vazio("({}) {} ({}) {}".format(ddd1, telefone1, ddd2, telefone2))
            sheet.cell(row=linha_planilha, column=6).value = email
            #sheet.cell(row=linha_planilha, column=7).value = remove_caracteres(capital_social)
            #sheet.cell(row=linha_planilha, column=8).value = ', '.join(socios)
            
            wb.save('AiimsColetados2.xlsx')
            
            messagebox.showinfo("Sucesso", "Informações salvas com sucesso!")
        else:
            messagebox.showerror("Erro", f"Erro ao acessar o CNPJ: {requisicao.status_code}")
    else:
        messagebox.showerror("CNPJ Inválido", "O CNPJ deve conter 14 números.")

def remove_caracter(string):
    nova_string = string.replace(".", " ").replace("'", " ").replace("-", " ").replace("/", " ").replace("_", " ").replace("&", " ")
    return nova_string

def remove_caracteres(string):
    nova_string = string.replace(".", "").replace(",", "")
    return nova_string

def vazio(string):
    nova_string = string.replace("None", " ")
    return nova_string

def cnpj(string):
    nova_string = string.replace(" ", "").replace("-", "").replace(".", "").replace("/", "").replace(" ", "")
    return nova_string
# Configurando a interface gráfica
root = tk.Tk()
root.title("Consulta de CNPJ")

frame = tk.Frame(root)
frame.pack(padx=20, pady=20)

label_cnpj = tk.Label(frame, text="Digite o CNPJ:")
label_cnpj.grid(row=0, column=0)

entry_cnpj = tk.Entry(frame)
entry_cnpj.grid(row=0, column=1)

button_consultar = tk.Button(frame, text="Consultar", command=consultar_cnpj)
button_consultar.grid(row=1, column=0, columnspan=2, pady=10)

button_sair = tk.Button(frame, text="Sair", command=root.destroy)
button_sair.grid(row=2, column=0, columnspan=2)

root.mainloop()